# openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
import openpyxl

inv_file = openpyxl.open("inventory.xlsx")

product_list = inv_file["Sheet1"]
# declare a dict, key = company name, value = products_num
products_per_supplier = {}
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):

    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    inventory_price = product_list.cell(product_row, 5)

    # calculation Of number of products per company
    # products_per_supplier是否有这个key
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        print("add a new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation of total value of inventory per company
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name) + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # add inventory price to excel
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
# save and create a new file
inv_file.save("inventory_with_total_value.xlsx")